# 1.
# У файлі task1.txt знаходиться текст субтитрів взятий з відео на ютубі. Текст
# складається з міток часу і репліки яка була сказана в той момент часу.
# Причому репліка знаходиться в наступному рядку після мітки часу.
# Прочитайте вміст файлу
# Результатом виконнання завдання повинно бути:
# 1. словник елементами якого буде пара ключ:значення де ключ - мітка часу,
# значення - репліка в даний момент часу
# 2. файл в якому знаходиться текст з якого видалені всі мітки часу. всі
# субтитри повинні мати вигляд простого тексту.
# Це означає що окрім видалення міток часу, вам потрібно видалити переноси
# рядків.
<<<<<<< HEAD

with open('task1.txt', 'rt') as file_1:
    list_comp = file_1.readlines()
    # print(list_comp)
    list_comp = [block.strip() for block in list_comp]
    dict_comp = dict(zip(list_comp[::2], list_comp[1::2]))
    print(dict_comp)

with open('task1_new.txt', 'wt') as file_2:
    for value in dict_comp:
        file_2.write(dict_comp[value])

# Now file will be to read only subtitles without time and margin.

with open('task1_new.txt', 'rt+') as file_2:
    print(file_2.read())


# 2.
# В файлі task2 збережений список, відкрийте цей файл, прочитайте вміст, і
# знайдіть середнє арифметичне чисел що знаходяться в списку.

from pickle import load
=======

with open('task1.txt', 'rt') as file_1:
    list_comp = file_1.readlines()
    # print(list_comp)
    list_comp = [block.strip() for block in list_comp]
    dict_comp = dict(zip(list_comp[::2], list_comp[1::2]))
    print(dict_comp)

with open('task1_new.txt', 'wt') as file_2:
    for value in dict_comp:
        file_2.write(dict_comp[value])

# Now file will be to read only subtitles without time and margin.

with open('task1_new.txt', 'rt+') as file_2:
    print(file_2.read())


# 2.
# В файлі task2 збережений список, відкрийте цей файл, прочитайте вміст, і
# знайдіть середнє арифметичне чисел що знаходяться в списку.

from pickle import load

>>>>>>> 46648fa0f9c444b443f32ebc1689c7932251c9a1

with open('task2', 'rb') as file_2:
    return_file_2 = load(file_2)
    print(return_file_2)

average_sum = sum(return_file_2) / len(return_file_2)
print("The average sum this number equal:", average_sum)


# 3. Використовуючи openpyxl (або будь-яку іншу зручну для вас бібліотеку),
# напишіть контекстний менеджер для роботи з ексель.
# Даний менеджер повинен бути аналогом методу open()
from openpyxl import Workbook
import datetime

wb = Workbook()

class Task:
    def __init__(self, path):
        self.file = open(path)
    def __enter__(self):
        return self.file
    def __exit__(self, exc_type, exc_val, exc_tb):
        if exc_type is None:
            return True
        elif exc_val is Exception:
            print("Exception")
            return True
        else:
            return False
task = Task

with open('task_3.xlsx', 'r') as file:
    # grab the active worksheet
    ws = wb.active

# Data can be assigned directly to cells
    ws['A1'] = 'Performed a three task.'

# Rows can also be appended
    file.write = ws.append([1])
    file.write = ws.append([2])
    file.write = ws.append([3])

# Python types will automatically be converted
    ws['A7'] = datetime.datetime.now()

# Save the file
    wb.save("task_3.xlsx")
    print(file)